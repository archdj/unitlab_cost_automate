"""F4 업로드 로더 — CSV / Excel / Notion ExportBlock zip.

각 파일을 읽어서 표준 컬럼 (excel_construction.md §1) 으로 정규화한 행 목록 + 검증 리포트를 반환한다.
운영 DB 에는 쓰지 않는다 (read-only). 결과는 호출자가 `harness/reports/uploads/` 에 저장하거나 미리보기로만 사용.
"""
from __future__ import annotations

import csv
import io
import re
import zipfile
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import IO

import openpyxl

# ─────────────────────────────────────────────────────────────────────────────
# 헤더 자동 인식 키워드 (excel_construction.md §2)
# ─────────────────────────────────────────────────────────────────────────────

HEADER_KEYWORDS = {
    "raw_description":  ["자재", "품목", "품명", "자재명", "부재명", "item", "name", "material"],
    "actual_quantity":  ["수량", "물량", "qty", "quantity"],
    "unit":             ["단위", "unit"],
    "unit_price":       ["단가", "price", "unit_price"],
    "total_amount":     ["금액", "공급가", "supply", "total", "합계", "amount"],
    "work_code_raw":    ["공종", "분류", "work", "category"],
    "vendor_name":      ["업체", "거래처", "vendor", "supplier", "업체명"],
    "settlement_date":  ["결제일", "정산일", "견적일", "납품일", "일자", "date"],
    "invoice_no":       ["세금계산서", "청구서", "invoice", "계산서"],
    "project_name":     ["프로젝트", "프로젝트명", "현장", "project"],
    "spec":             ["규격", "형상", "spec", "size"],
}

REQUIRED_FIELDS = ("raw_description",)
AT_LEAST_ONE_OF = ("total_amount", "unit_price", "actual_quantity")  # 금액 산정 단서

# ─────────────────────────────────────────────────────────────────────────────
# 결과 모델
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class UploadIssue:
    severity: str  # "block" | "warn" | "info"
    code: str      # "missing_required" 등
    message: str
    row: int | None = None


@dataclass
class UploadReport:
    source_filename: str
    source_type: str          # "csv" | "xlsx" | "notion_zip"
    sheet_name: str | None
    total_rows: int
    accepted_rows: int
    rejected_rows: int
    column_mapping: dict[str, str | None]   # 원본 헤더 → 표준 필드
    unmapped_columns: list[str]
    preview: list[dict]                     # 정규화 후 처음 N행
    issues: list[UploadIssue] = field(default_factory=list)
    sheet_choices: list[str] | None = None  # xlsx 일 때
    alias_stats: dict | None = None         # raw_description ↔ material_aliases 매칭 통계
    saved_to: str | None = None             # 디스크 저장 경로 (commit 일 때)

    def to_dict(self) -> dict:
        return {
            "source_filename": self.source_filename,
            "source_type":     self.source_type,
            "sheet_name":      self.sheet_name,
            "sheet_choices":   self.sheet_choices,
            "total_rows":      self.total_rows,
            "accepted_rows":   self.accepted_rows,
            "rejected_rows":   self.rejected_rows,
            "column_mapping":  self.column_mapping,
            "unmapped_columns": self.unmapped_columns,
            "preview":         self.preview,
            "issues":          [i.__dict__ for i in self.issues],
            "alias_stats":     self.alias_stats,
            "saved_to":        self.saved_to,
        }


# ─────────────────────────────────────────────────────────────────────────────
# 헤더 매핑 / 값 파서
# ─────────────────────────────────────────────────────────────────────────────

def map_header(headers: list[str]) -> tuple[dict[str, str | None], list[str]]:
    """원본 헤더(원어) → 표준 필드. 정확 → prefix/suffix → 부분 매칭 순으로 우선순위.

    예: "수량" 키워드는 "수량" 헤더와 정확 매칭되면, "수량산출서" 같은 부분 매칭보다 우선.
    """
    by_lower = {h: (h or "").strip().lower() for h in headers}
    mapping: dict[str, str | None] = {std: None for std in HEADER_KEYWORDS}
    used: set[str] = set()

    def try_match(predicate):
        for std, kws in HEADER_KEYWORDS.items():
            if mapping[std]:
                continue
            for h in headers:
                if not h or h in used:
                    continue
                low = by_lower[h]
                if any(predicate(low, kw.lower()) for kw in kws):
                    mapping[std] = h
                    used.add(h)
                    break

    try_match(lambda low, kw: low == kw)                       # 1) 정확
    try_match(lambda low, kw: low.startswith(kw) or low.endswith(kw))  # 2) prefix/suffix
    try_match(lambda low, kw: kw in low)                       # 3) 부분 (fallback)

    unmapped = [h for h in headers if h and h not in used]
    return mapping, unmapped


_NUM_RE = re.compile(r"[^\d\.\-]")


def parse_number(v) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    cleaned = _NUM_RE.sub("", s)
    if cleaned in ("", ".", "-", "-."):
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


_DATE_FORMATS = [
    "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d",
    "%Y. %m. %d",
    "%Y%m%d",
]


def parse_date(v) -> str | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    s = str(v).strip()
    if not s:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def attach_alias_match(rows: list[dict], alias_idx: dict | None) -> dict:
    """행마다 alias 매칭을 시도해서 결과를 행에 추가하고 통계 반환."""
    from .db import match_alias  # circular 회피
    if alias_idx is None:
        return {"attempted": 0, "matched_exact": 0, "matched_token": 0}
    exact = token = 0
    for row in rows:
        raw = row.get("raw_description")
        m = match_alias(raw or "", alias_idx)
        row["match_method"]   = m["method"]
        row["matched_alias"]  = m["alias_name"]
        row["matched_material_id"] = m["material_id"]
        row["match_score"]    = m["score"]
        if m["method"] == "exact":
            exact += 1
        elif m["method"] == "token":
            token += 1
    return {
        "attempted":     len(rows),
        "matched_exact": exact,
        "matched_token": token,
        "match_rate":    round((exact + token) / max(len(rows), 1), 4),
    }


def normalize_row(row: dict, mapping: dict[str, str | None]) -> tuple[dict, list[UploadIssue]]:
    issues: list[UploadIssue] = []
    out: dict = {}

    def get(field: str):
        col = mapping.get(field)
        return row.get(col) if col else None

    out["raw_description"] = (str(get("raw_description") or "")).strip() or None
    out["actual_quantity"] = parse_number(get("actual_quantity"))
    out["unit"]            = (str(get("unit") or "")).strip() or None
    out["unit_price"]      = parse_number(get("unit_price"))
    out["total_amount"]    = parse_number(get("total_amount"))
    out["work_code_raw"]   = (str(get("work_code_raw") or "")).strip() or None
    out["vendor_name"]     = (str(get("vendor_name") or "")).strip() or None
    out["settlement_date"] = parse_date(get("settlement_date"))
    out["invoice_no"]      = (str(get("invoice_no") or "")).strip() or None
    out["project_name"]    = (str(get("project_name") or "")).strip() or None
    out["spec"]            = (str(get("spec") or "")).strip() or None

    # total_amount 가 비고 unit_price * quantity 가능하면 보강
    if out["total_amount"] is None and out["unit_price"] and out["actual_quantity"]:
        out["total_amount"] = out["unit_price"] * out["actual_quantity"]
    # 반대 케이스: total + qty 로 unit_price 역산
    elif out["unit_price"] is None and out["total_amount"] and out["actual_quantity"]:
        out["unit_price"] = out["total_amount"] / out["actual_quantity"]

    # 검증
    for field in REQUIRED_FIELDS:
        if not out.get(field):
            issues.append(UploadIssue("block", "missing_required",
                                      f"필수 필드 '{field}' 누락"))
    if all(out.get(f) is None for f in AT_LEAST_ONE_OF):
        issues.append(UploadIssue("block", "no_amount_signal",
                                  "total_amount/unit_price/actual_quantity 모두 비어있음"))
    if out["total_amount"] is not None and out["total_amount"] < 0:
        issues.append(UploadIssue("warn", "negative_amount", "음수 금액"))
    if (out["unit_price"] and out["actual_quantity"] and out["total_amount"]
            and abs(out["unit_price"] * out["actual_quantity"] - out["total_amount"])
                / max(out["total_amount"], 1) > 0.01):
        issues.append(UploadIssue("warn", "amount_mismatch",
                                  "unit_price × quantity ≠ total_amount (오차 1% 초과)"))

    return out, issues


# ─────────────────────────────────────────────────────────────────────────────
# 형식별 로더
# ─────────────────────────────────────────────────────────────────────────────

def load_csv(stream: IO[bytes], filename: str, *, preview_rows: int = 5) -> UploadReport:
    text = stream.read()
    # 인코딩 후보 순회
    decoded = None
    for enc in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            decoded = text.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if decoded is None:
        return UploadReport(filename, "csv", None, 0, 0, 0, {}, [], [], [
            UploadIssue("block", "decode_failed", "UTF-8 / CP949 모두 디코드 실패")
        ])

    rdr = csv.reader(io.StringIO(decoded))
    rows = list(rdr)
    if not rows:
        return UploadReport(filename, "csv", None, 0, 0, 0, {}, [], [])

    headers = [h.strip() for h in rows[0]]
    mapping, unmapped = map_header(headers)
    return _finalize_table(filename, "csv", None, headers, rows[1:], mapping, unmapped, preview_rows)


def load_xlsx(stream: IO[bytes], filename: str, *, sheet: str | None = None,
              preview_rows: int = 5) -> UploadReport:
    wb = openpyxl.load_workbook(io.BytesIO(stream.read()), read_only=True, data_only=True)
    sheet_choices = wb.sheetnames
    chosen = sheet or sheet_choices[0]
    if chosen not in sheet_choices:
        return UploadReport(filename, "xlsx", chosen, 0, 0, 0, {}, [], [],
                            [UploadIssue("block", "sheet_not_found",
                                         f"시트 '{chosen}' 없음")],
                            sheet_choices=sheet_choices)
    ws = wb[chosen]
    grid = [[c for c in r] for r in ws.iter_rows(values_only=True)]

    # 헤더 행 자동 탐지: 비숫자 셀이 가장 많은 행 (상위 30행 중)
    header_idx = _detect_header_row(grid)
    if header_idx is None:
        return UploadReport(filename, "xlsx", chosen, 0, 0, 0, {}, [], [],
                            [UploadIssue("block", "header_not_detected",
                                         "헤더 행을 찾지 못함")],
                            sheet_choices=sheet_choices)

    headers = [_to_str(c) for c in grid[header_idx]]
    mapping, unmapped = map_header(headers)
    body = grid[header_idx + 1:]
    rep = _finalize_table(filename, "xlsx", chosen, headers, body, mapping, unmapped, preview_rows)
    rep.sheet_choices = sheet_choices
    rep.issues.append(UploadIssue("info", "header_row_detected",
                                  f"헤더를 {header_idx + 1}행으로 인식"))
    return rep


def load_notion_zip(stream: IO[bytes], filename: str, *, preview_rows: int = 5) -> UploadReport:
    """노션 ExportBlock zip — 중첩 zip 자동 추출 후 첫 CSV 처리."""
    raw = stream.read()
    csv_bytes, inner_name = _extract_first_csv(raw)
    if csv_bytes is None:
        return UploadReport(filename, "notion_zip", None, 0, 0, 0, {}, [], [], [
            UploadIssue("block", "no_csv_in_zip", "zip 안에서 CSV 를 찾지 못함")
        ])
    rep = load_csv(io.BytesIO(csv_bytes), inner_name or filename, preview_rows=preview_rows)
    rep.source_type = "notion_zip"
    rep.source_filename = f"{filename} → {inner_name}"
    return rep


def _extract_first_csv(blob: bytes, depth: int = 0) -> tuple[bytes | None, str | None]:
    if depth > 4:
        return None, None
    try:
        z = zipfile.ZipFile(io.BytesIO(blob))
    except zipfile.BadZipFile:
        return None, None
    for name in z.namelist():
        if name.lower().endswith(".csv"):
            return z.read(name), name
    # CSV 없으면 안의 첫 zip 을 다시 풀어본다 (중첩)
    for name in z.namelist():
        if name.lower().endswith(".zip"):
            return _extract_first_csv(z.read(name), depth + 1)
    return None, None


# ─────────────────────────────────────────────────────────────────────────────
# 헤더 행 탐지 / 정규화 후 마무리
# ─────────────────────────────────────────────────────────────────────────────

def _to_str(c) -> str:
    return "" if c is None else str(c).strip()


def _detect_header_row(grid: list[list], scan: int = 30) -> int | None:
    best_idx, best_score = None, 0
    for i, row in enumerate(grid[:scan]):
        cells = [c for c in row if _to_str(c)]
        # 비숫자 비율이 높을수록 헤더 후보
        score = sum(1 for c in cells if not isinstance(c, (int, float)))
        # 매칭되는 표준 키워드가 있을 때 가산점
        bonus = sum(1 for c in cells for std, kws in HEADER_KEYWORDS.items()
                    if any(kw.lower() in _to_str(c).lower() for kw in kws))
        total = score + bonus * 2
        if total > best_score and bonus >= 2:
            best_idx, best_score = i, total
    return best_idx


def _finalize_table(
    filename: str,
    src_type: str,
    sheet: str | None,
    headers: list[str],
    body: list[list],
    mapping: dict[str, str | None],
    unmapped: list[str],
    preview_n: int,
) -> UploadReport:
    accepted, rejected = 0, 0
    preview: list[dict] = []
    issues: list[UploadIssue] = []
    total = 0

    for i, row in enumerate(body, start=1):
        if not any(_to_str(c) for c in row):
            continue
        total += 1
        rec = {headers[j]: row[j] if j < len(row) else None for j in range(len(headers))}
        norm, row_issues = normalize_row(rec, mapping)
        is_block = any(x.severity == "block" for x in row_issues)
        if is_block:
            rejected += 1
        else:
            accepted += 1
        for x in row_issues:
            x.row = i
            issues.append(x)
        if len(preview) < preview_n:
            preview.append({"row": i, **norm})

    return UploadReport(
        source_filename=filename,
        source_type=src_type,
        sheet_name=sheet,
        total_rows=total,
        accepted_rows=accepted,
        rejected_rows=rejected,
        column_mapping=mapping,
        unmapped_columns=unmapped,
        preview=preview,
        issues=issues[:200],   # 너무 많으면 절단
    )


# ─────────────────────────────────────────────────────────────────────────────
# Dispatch
# ─────────────────────────────────────────────────────────────────────────────

def load_any(stream: IO[bytes], filename: str, *, sheet: str | None = None,
             preview_rows: int = 5) -> UploadReport:
    ext = Path(filename).suffix.lower()
    if ext == ".csv":
        return load_csv(stream, filename, preview_rows=preview_rows)
    if ext in (".xlsx", ".xlsm"):
        return load_xlsx(stream, filename, sheet=sheet, preview_rows=preview_rows)
    if ext == ".zip":
        return load_notion_zip(stream, filename, preview_rows=preview_rows)
    return UploadReport(filename, ext.lstrip(".") or "unknown", None, 0, 0, 0, {}, [], [], [
        UploadIssue("block", "unsupported_extension",
                    f"지원하지 않는 확장자: {ext}")
    ])


def enrich_with_alias_match(report: UploadReport, alias_idx: dict | None) -> None:
    """report.preview 의 행에 alias 매칭 추가하고 통계 기록."""
    if not report.preview:
        return
    report.alias_stats = attach_alias_match(report.preview, alias_idx)
